import csv
import io
import logging
from io import BytesIO
from datetime import datetime, timedelta
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.utils.translation import gettext as _
from django.conf import settings

from .models import InvitedGuest, GuestResponse

logger = logging.getLogger(__name__)

# Tentative d'import de reportlab (optionnel)
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm, cm
    import qrcode
    from PIL import Image as PILImage
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logger.warning("reportlab not installed. PDF generation disabled.")


def import_guests_from_excel(excel_file, event, user):
    import openpyxl
    from apps.events.models import Table

    result = {'created': 0, 'errors': 0, 'error_messages': []}

    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        # Déterminer les colonnes (on suppose que l'en-tête est sur la ligne 1)
        headers = [cell.value for cell in ws[1] if cell.value]
        # On cherche les colonnes par nom (flexible)
        col_map = {}
        for idx, header in enumerate(headers):
            if header and isinstance(header, str):
                header_lower = header.lower().strip()
                if 'prénom' in header_lower or 'prenom' in header_lower:
                    col_map['first_name'] = idx
                elif 'nom' in header_lower and 'post' not in header_lower:
                    col_map['last_name'] = idx
                elif 'postnom' in header_lower or 'post-nom' in header_lower:
                    col_map['middle_name'] = idx
                elif 'email' in header_lower:
                    col_map['email'] = idx
                elif 'téléphone' in header_lower or 'phone' in header_lower:
                    col_map['phone'] = idx
                elif 'table' in header_lower:
                    col_map['table'] = idx

        # Si on n'a pas de colonne table, on ignore
        has_table_col = 'table' in col_map

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue

            def get_cell(col_name):
                idx = col_map.get(col_name)
                if idx is not None and idx < len(row):
                    val = row[idx]
                    return str(val).strip() if val else ''
                return ''

            first_name = get_cell('first_name')
            last_name = get_cell('last_name')
            middle_name = get_cell('middle_name')
            email = get_cell('email') or None
            phone = get_cell('phone')
            table_name = get_cell('table') if has_table_col else ''

            if not first_name or not last_name:
                result['errors'] += 1
                result['error_messages'].append(f"Ligne {row_idx}: Prénom ou nom manquant")
                continue

            # Gérer la table : la créer si elle n'existe pas
            table = None
            if table_name:
                table, _ = Table.objects.get_or_create(
                    event=event,
                    number=table_name,
                    defaults={'name': f"Table {table_name}", 'capacity': 10}  # capacité par défaut
                )

            try:
                invited_guest, created = InvitedGuest.objects.get_or_create(
                    event=event,
                    email=email,
                    defaults={
                        'first_name': first_name,
                        'last_name': last_name,
                        'middle_name': middle_name,
                        'phone': phone,
                        'created_by': user,
                        'table': table,  # assignation
                    }
                )
                if not created:
                    # Mise à jour
                    invited_guest.first_name = first_name
                    invited_guest.last_name = last_name
                    invited_guest.middle_name = middle_name
                    invited_guest.phone = phone
                    invited_guest.table = table
                    invited_guest.save()
                    result['updated'] += 1
                else:
                    result['created'] += 1
            except Exception as e:
                result['errors'] += 1
                result['error_messages'].append(f"Ligne {row_idx}: {str(e)}")

    except Exception as e:
        result['errors'] += 1
        result['error_messages'].append(f"Erreur lecture fichier: {str(e)}")

    return result

def generate_invitation_pdf(guest_response):
    """
    Génère un PDF d'invitation électronique.
    Si reportlab n'est pas installé, retourne un message d'erreur.
    """
    if not REPORTLAB_AVAILABLE:
        logger.error("reportlab is not installed. Cannot generate PDF.")
        return None
    
    event = guest_response.event
    buffer = BytesIO()
    
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
        from reportlab.lib.units import mm
        import qrcode
        from PIL import Image as PILImage
        
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=72)
        styles = getSampleStyleSheet()
        story = []
        
        # Style personnalisé
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2C3E50'),
            alignment=1,
            spaceAfter=20
        )
        
        # Titre
        story.append(Paragraph(f"Invitation - {event.name}", title_style))
        story.append(Spacer(1, 10))
        story.append(Paragraph(f"Pour: {guest_response.get_full_name()}", title_style))
        story.append(Spacer(1, 20))
        
        # Informations
        info_data = [
            ["📅 Date:", event.date.strftime('%d %B %Y') if event.date else 'À confirmer'],
            ["⏰ Heure:", event.time.strftime('%H:%M') if event.time else 'À confirmer'],
            ["📍 Lieu:", event.location],
        ]
        if event.google_maps_link:
            info_data.append(["🗺️ Maps:", f'<link href="{event.google_maps_link}">Voir sur Google Maps</link>'])
        
        info_table = Table(info_data, colWidths=[80, 350])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 20))
        
        if event.dress_code:
            story.append(Paragraph(f"👔 Code vestimentaire: {event.dress_code}", styles['Normal']))
            story.append(Spacer(1, 10))
        
        story.append(Paragraph("Nous sommes ravis de vous compter parmi nos invités !", styles['Normal']))
        story.append(Spacer(1, 10))
        story.append(Paragraph("Veuillez confirmer votre présence via le lien reçu par email.", styles['Normal']))
        story.append(Spacer(1, 30))
        
        # QR Code
        qr = qrcode.QRCode(version=1, box_size=8, border=2)
        qr.add_data(guest_response.get_invitation_link())
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")
        qr_buffer = BytesIO()
        qr_img.save(qr_buffer, format='PNG')
        qr_buffer.seek(0)
        
        qr_pil = PILImage.open(qr_buffer)
        from reportlab.lib.utils import ImageReader
        qr_reader = ImageReader(qr_pil)
        
        qr_table = Table([[Image(qr_reader, width=100*mm, height=100*mm)]])
        qr_table.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'CENTER')]))
        story.append(qr_table)
        story.append(Spacer(1, 20))
        story.append(Paragraph("Scannez ce QR code pour confirmer votre présence", styles['Italic']))
        
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    except Exception as e:
        logger.error(f"Erreur génération PDF: {str(e)}")
        return None


def send_reminders_for_event(event, days_before=7):
    """
    Envoie des rappels aux invités qui n'ont pas encore répondu.
    """
    if not event.date:
        return 0
    
    today = datetime.now().date()
    event_date = event.date
    days_until = (event_date - today).days
    
    if days_until != days_before:
        return 0
    
    invited_guests = event.invited_guests.all()
    responded_emails = event.responses.filter(
        will_attend=True,
        verification_status='verified'
    ).values_list('email', flat=True)
    
    pending_guests = [g for g in invited_guests if g.email and g.email not in responded_emails]
    
    sent_count = 0
    for guest in pending_guests:
        try:
            response, created = GuestResponse.objects.get_or_create(
                event=event,
                email=guest.email,
                defaults={
                    'first_name': guest.first_name,
                    'last_name': guest.last_name,
                }
            )
            if response.send_reminder():
                sent_count += 1
        except Exception as e:
            logger.error(f"Erreur rappel pour {guest.email}: {str(e)}")
    
    return sent_count

class TableAssignmentService:
    def __init__(self, event):
        self.event = event

    def auto_assign_all(self):
        tables = list(self.event.tables.all().order_by('number'))
        if not tables:
            return False
        # Récupérer les invités qui n'ont pas encore de table
        guests = list(self.event.responses.filter(will_attend=True, table__isnull=True))
        if not guests:
            return False
        for guest in guests:
            for table in tables:
                if table.guests.count() < table.capacity:
                    guest.table = table
                    guest.save()
                    break
        return True
    
def generate_tables_pdf(event):
    """Génère un PDF avec la liste des tables, invités et leurs boissons."""
    if not REPORTLAB_AVAILABLE:
        return None
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    story.append(Paragraph(f"Récapitulatif des tables - {event.name}", styles['Title']))
    story.append(Spacer(1, 12))
    story.append(Paragraph(f"Date : {event.date.strftime('%d/%m/%Y')} à {event.time.strftime('%H:%M') if event.time else ''}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    tables = event.tables.all().order_by('number')
    for table in tables:
        story.append(Paragraph(f"Table {table.number} (capacité {table.capacity})", styles['Heading2']))
        guests = table.guests.filter(will_attend=True).order_by('last_name', 'first_name')
        if guests:
            data = [["Prénom", "Nom", "Boisson"]]
            for g in guests:
                data.append([g.first_name, g.last_name, g.drink_display])
            t = Table(data, colWidths=[100,100,150])
            t.setStyle(TableStyle([
                ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
                ('FONTSIZE', (0,0), (-1,-1), 10),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ]))
            story.append(t)
        else:
            story.append(Paragraph("Aucun invité confirmé sur cette table.", styles['Italic']))
        story.append(Spacer(1, 12))
    
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()

import io
from django.core.files.base import ContentFile
from django.conf import settings
import qrcode
from PIL import Image as PILImage

logger = logging.getLogger(__name__)

# Tentative d'import de reportlab
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm, cm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.fonts import addMapping
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logger.warning("reportlab not installed. PDF generation disabled.")


class InvitationPDFService:
    """
    Service de génération d'invitation PDF avec QR code.
    """
    
    def __init__(self, guest_response):
        self.guest_response = guest_response
        self.event = guest_response.event
    
    def generate(self):
        """
        Génère le PDF d'invitation et retourne le contenu binaire.
        """
        if not REPORTLAB_AVAILABLE:
            logger.error("reportlab is not installed. Cannot generate PDF.")
            return None
        
        buffer = BytesIO()
        
        try:
            # Création du document
            doc = SimpleDocTemplate(
                buffer,
                pagesize=A4,
                rightMargin=50,
                leftMargin=50,
                topMargin=50,
                bottomMargin=50
            )
            
            styles = getSampleStyleSheet()
            story = []
            
            # Styles personnalisés
            title_style = ParagraphStyle(
                'TitleStyle',
                parent=styles['Heading1'],
                fontSize=28,
                textColor=colors.HexColor('#2C3E50'),
                alignment=TA_CENTER,
                spaceAfter=30,
                fontName='Helvetica-Bold'
            )
            
            subtitle_style = ParagraphStyle(
                'SubtitleStyle',
                parent=styles['Heading2'],
                fontSize=18,
                textColor=colors.HexColor('#8B5CF6'),
                alignment=TA_CENTER,
                spaceAfter=20,
                fontName='Helvetica'
            )
            
            body_style = ParagraphStyle(
                'BodyStyle',
                parent=styles['Normal'],
                fontSize=12,
                textColor=colors.HexColor('#333333'),
                alignment=TA_LEFT,
                spaceAfter=8,
                fontName='Helvetica'
            )
            
            # --- Titre ---
            story.append(Paragraph("INVITATION", title_style))
            story.append(Spacer(1, 10))
            story.append(Paragraph(self.event.name, subtitle_style))
            story.append(Spacer(1, 20))
            
            # --- Destinataire ---
            guest_name = self.guest_response.get_full_name()
            salutation = "Madame" if self.guest_response.first_name.lower() in [
                'marie', 'jeanne', 'claire', 'sophie', 'anne', 'catherine', 
                'elisabeth', 'françoise', 'nathalie', 'isabelle', 'sylvie', 
                'christine', 'laurence', 'dominique', 'valérie', 'patricia', 
                'brigitte', 'nicole', 'monique', 'micheline', 'marguerite', 
                'yolande', 'germaine', 'léontine', 'joséphine', 'mariette'
            ] else "Monsieur"
            
            story.append(Paragraph(
                f"{salutation} <b>{guest_name}</b>",
                body_style
            ))
            story.append(Spacer(1, 10))
            
            # --- Corps de l'invitation ---
            story.append(Paragraph(
                "Nous avons le plaisir de vous inviter à notre événement :",
                body_style
            ))
            story.append(Spacer(1, 15))
            
            # Détails de l'événement
            details_data = [
                ["Date :", self.event.date.strftime('%d %B %Y') if self.event.date else 'À confirmer'],
                ["Heure :", self.event.time.strftime('%H:%M') if self.event.time else 'À confirmer'],
                ["Lieu :", self.event.location],
            ]
            if self.event.dress_code:
                details_data.append(["Tenue :", self.event.dress_code])
            
            details_table = Table(details_data, colWidths=[80, 350])
            details_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 12),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E0E0E0')),
            ]))
            story.append(details_table)
            story.append(Spacer(1, 20))
            
            # --- Table assignée ---
            if self.guest_response.table:
                table_style = ParagraphStyle(
                    'TableStyle',
                    parent=styles['Normal'],
                    fontSize=14,
                    textColor=colors.HexColor('#8B5CF6'),
                    alignment=TA_CENTER,
                    fontName='Helvetica-Bold'
                )
                story.append(Paragraph(f"Table assignée : {self.guest_response.table.number}", table_style))
                story.append(Spacer(1, 15))
            
            # --- Message de fin ---
            story.append(Paragraph(
                "Nous sommes impatients de vous accueillir et de partager ce moment avec vous.",
                body_style
            ))
            story.append(Spacer(1, 10))
            story.append(Paragraph(
                "Veuillez confirmer votre présence en scannant le QR code ci-dessous.",
                body_style
            ))
            story.append(Spacer(1, 20))
            
            # --- QR Code ---
            qr = qrcode.QRCode(
                version=1,
                box_size=6,
                border=2
            )
            qr.add_data(self.guest_response.get_invitation_link())
            qr.make(fit=True)
            qr_img = qr.make_image(fill_color="black", back_color="white")
            
            qr_buffer = BytesIO()
            qr_img.save(qr_buffer, format='PNG')
            qr_buffer.seek(0)
            
            qr_pil = PILImage.open(qr_buffer)
            from reportlab.lib.utils import ImageReader
            qr_reader = ImageReader(qr_pil)
            
            qr_table = Table([[Image(qr_reader, width=60*mm, height=60*mm)]])
            qr_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'CENTER'),
            ]))
            story.append(qr_table)
            story.append(Spacer(1, 10))
            
            qr_text_style = ParagraphStyle(
                'QRTextStyle',
                parent=styles['Italic'],
                fontSize=10,
                textColor=colors.HexColor('#666666'),
                alignment=TA_CENTER
            )
            story.append(Paragraph("Scannez ce QR code pour confirmer votre présence", qr_text_style))
            
            # --- Pied de page ---
            story.append(Spacer(1, 30))
            footer_style = ParagraphStyle(
                'FooterStyle',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.HexColor('#999999'),
                alignment=TA_CENTER
            )
            year = self.event.date.strftime('%Y') if self.event.date else ''
            story.append(Paragraph(
                f"© {year} {self.event.name} - Document valable uniquement avec le QR code",
                footer_style
            ))
            
            # Construction du PDF
            doc.build(story)
            buffer.seek(0)
            
            return buffer.getvalue()
        
        except Exception as e:
            logger.error(f"Erreur génération PDF: {str(e)}")
            return None


def generate_invitation_pdf(guest_response):
    """
    Fonction d'export pour compatibilité avec le code existant.
    """
    service = InvitationPDFService(guest_response)
    return service.generate()