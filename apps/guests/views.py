import csv
import logging
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.utils.translation import gettext as _
from django.views import View
from django.views.generic import ListView, TemplateView, FormView, CreateView, DetailView, UpdateView, DeleteView
from django.utils import timezone
from django.urls import reverse, reverse_lazy
from django.conf import settings
from django.core.exceptions import PermissionDenied
from django.db.models import Q

from apps.events.models import Event, Table
from .models import GuestResponse, InvitedGuest
from .forms import RSVPForm, InvitedGuestForm, GuestBulkImportForm
from .services import import_guests_from_excel, generate_invitation_pdf

logger = logging.getLogger(__name__)


# ============================================================
# 1. LISTE DES RÉPONSES (GUEST RESPONSE)
# ============================================================

class GuestListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """Liste des réponses des invités pour un événement."""
    template_name = 'guests/guest_list.html'
    context_object_name = 'guests'

    def test_func(self):
        self.event = get_object_or_404(Event, id=self.kwargs.get('event_id'))
        user = self.request.user
        return (self.event.main_organizer == user or
                self.event.collaborators.filter(user=user, status='accepted').exists())

    def get_paginate_by(self, queryset):
        per_page = self.request.GET.get('per_page')
        if per_page and per_page.isdigit():
            per_page = int(per_page)
            if per_page in [10, 15, 20, 30, 50, 100]:
                return per_page
        return getattr(settings, 'GUESTS_PER_PAGE', 20)

    def get_queryset(self):
        return self.event.responses.all().order_by('-submitted_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['event'] = self.event
        context['current_per_page'] = self.get_paginate_by(self.get_queryset())
        responses = self.event.responses
        context['stats'] = {
            'total': responses.count(),
            'attending': responses.filter(will_attend=True).count(),
            'not_attending': responses.filter(will_attend=False).count(),
            'verified': responses.filter(verification_status='verified').count(),
            'unverified': responses.filter(verification_status='unverified').count(),
        }
        return context


# ============================================================
# 2. EXPORTS DES RÉPONSES (CSV / EXCEL)
# ============================================================

class ExportGuestsCSVView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Export CSV des réponses invités."""
    
    def test_func(self):
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        return (self.event.main_organizer == self.request.user or
                self.event.collaborators.filter(user=self.request.user, status='accepted').exists())

    def get(self, request, event_id):
        responses = self.event.responses.all().order_by('-submitted_at')
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="responses_{self.event.id}.csv"'
        response.write('\ufeff')
        
        writer = csv.writer(response)
        writer.writerow([
            _('First name'), _('Last name'), _('Email'), _('Phone'),
            _('Will attend'), _('Number of guests'), _('Accompanied'),
            _('Drink choice'), _('Other drink'), _('Verification status'),
            _('Submitted at')
        ])
        
        for r in responses:
            writer.writerow([
                r.first_name, r.last_name, r.email, r.phone,
                _('Yes') if r.will_attend else _('No'),
                r.number_of_guests,
                _('Yes') if r.is_accompanied else _('No'),
                r.drink_display, r.drink_other or '',
                r.get_verification_status_display(),
                r.submitted_at.strftime('%d/%m/%Y %H:%M'),
            ])
        return response


class ExportGuestsExcelView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Export Excel des réponses invités."""
    
    def test_func(self):
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        return (self.event.main_organizer == self.request.user or
                self.event.collaborators.filter(user=self.request.user, status='accepted').exists())

    def get(self, request, event_id):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        
        responses = self.event.responses.all().order_by('-submitted_at')
        
        wb = Workbook()
        ws = wb.active
        ws.title = str(_('Responses'))
        
        headers = [
            _('First name'), _('Last name'), _('Email'), _('Phone'),
            _('Will attend'), _('Number of guests'), _('Accompanied'),
            _('Drink choice'), _('Other drink'), _('Verification status'),
            _('Submitted at')
        ]
        
        header_font = Font(bold=True)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for row, r in enumerate(responses, 2):
            ws.cell(row=row, column=1, value=r.first_name)
            ws.cell(row=row, column=2, value=r.last_name)
            ws.cell(row=row, column=3, value=r.email)
            ws.cell(row=row, column=4, value=r.phone)
            ws.cell(row=row, column=5, value=_('Yes') if r.will_attend else _('No'))
            ws.cell(row=row, column=6, value=r.number_of_guests)
            ws.cell(row=row, column=7, value=_('Yes') if r.is_accompanied else _('No'))
            ws.cell(row=row, column=8, value=r.drink_display)
            ws.cell(row=row, column=9, value=r.drink_other or '')
            ws.cell(row=row, column=10, value=r.get_verification_status_display())
            ws.cell(row=row, column=11, value=r.submitted_at.strftime('%d/%m/%Y %H:%M'))
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="responses_{self.event.id}.xlsx"'
        wb.save(response)
        return response


# ============================================================
# 3. EXPORT DES CHECK-INS (H)
# ============================================================

class ExportCheckinsCSVView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Export CSV des check-ins."""
    
    def test_func(self):
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        return self.request.user == self.event.main_organizer

    def get(self, request, event_id):
        responses = GuestResponse.objects.filter(
            event=self.event,
            checkin_time__isnull=False
        ).select_related('table').order_by('checkin_time')
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="checkins_{self.event.id}.csv"'
        response.write('\ufeff')
        
        writer = csv.writer(response)
        writer.writerow(['Nom', 'Email', 'Table', "Heure d'arrivée"])
        
        for r in responses:
            writer.writerow([
                r.get_full_name(),
                r.email,
                r.table.number if r.table else '-',
                r.checkin_time.strftime('%d/%m/%Y %H:%M')
            ])
        return response


class ExportCheckinsExcelView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Export Excel des check-ins."""
    
    def test_func(self):
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        return self.request.user == self.event.main_organizer

    def get(self, request, event_id):
        from openpyxl import Workbook
        from openpyxl.styles import Font
        
        responses = GuestResponse.objects.filter(
            event=self.event,
            checkin_time__isnull=False
        ).select_related('table').order_by('checkin_time')
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Check-ins"
        
        headers = ['Nom', 'Email', 'Table', "Heure d'arrivée"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).font = Font(bold=True)
        
        for row, r in enumerate(responses, 2):
            ws.cell(row=row, column=1, value=r.get_full_name())
            ws.cell(row=row, column=2, value=r.email)
            ws.cell(row=row, column=3, value=r.table.number if r.table else '-')
            ws.cell(row=row, column=4, value=r.checkin_time.strftime('%d/%m/%Y %H:%M'))
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="checkins_{self.event.id}.xlsx"'
        wb.save(response)
        return response


# ============================================================
# 4. RSVP - MERCI
# ============================================================

class RSVPThanksView(TemplateView):
    """Page de remerciement après RSVP."""
    template_name = 'guests/rsvp_thanks.html'


# ============================================================
# 5. IMPORT EXCEL
# ============================================================

class BulkImportGuestsView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Import Excel d'invités pré-enregistrés."""
    template_name = 'guests/bulk_import.html'

    def test_func(self):
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        return self.request.user == self.event.main_organizer

    def get(self, request, event_id):
        form = GuestBulkImportForm()
        return render(request, self.template_name, {'form': form, 'event': self.event})

    def post(self, request, event_id):
        form = GuestBulkImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            result = import_guests_from_excel(excel_file, self.event, request.user)
            messages.success(
                request,
                _('Import terminé: %(created)s invités ajoutés, %(errors)s erreurs') % {
                    'created': result['created'],
                    'errors': result['errors']
                }
            )
            return redirect('guests:invited_list', event_id=self.event.id)
        return render(request, self.template_name, {'form': form, 'event': self.event})


# ============================================================
# 6. CHECK-IN
# ============================================================

class CheckInView(LoginRequiredMixin, UserPassesTestMixin, View):
    """
    Scanner un QR code (ou saisir un code court) et valider l'arrivée.
    """
    template_name = 'guests/checkin.html'
    success_template = 'guests/checkin_success.html'

    def test_func(self):
        return True  # Vérification manuelle dans dispatch

    def dispatch(self, request, *args, **kwargs):
        token = kwargs.get('token')
        try:
            import uuid
            uuid_obj = uuid.UUID(token)
            self.guest_response = get_object_or_404(GuestResponse, invitation_token=uuid_obj)
        except ValueError:
            self.guest_response = get_object_or_404(GuestResponse, short_code=token)

        user = request.user
        event = self.guest_response.event
        is_organizer = (event.main_organizer == user)
        is_collaborator_with_scan = event.collaborators.filter(user=user, can_scan=True).exists()

        if not (is_organizer or is_collaborator_with_scan):
            messages.error(request, _("Vous n'avez pas l'autorisation de scanner pour cet événement."))
            return redirect('events:event_list')

        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        guest = self.guest_response
        if guest.checkin_time:
            return render(request, self.template_name, {
                'guest': guest,
                'already_checked_in': True,
                'message': _("Cette invitation a déjà été scannée à {}.").format(
                    guest.checkin_time.strftime("%H:%M:%S")
                )
            })
        return render(request, self.template_name, {
            'guest': guest,
            'already_checked_in': False,
        })

    def post(self, request, *args, **kwargs):
        guest = self.guest_response
        if guest.checkin_time:
            messages.warning(request, _("Cette invitation a déjà été utilisée."))
            return redirect('guests:checkin', token=kwargs.get('token'))

        guest.checkin_time = timezone.now()
        guest.save(update_fields=['checkin_time'])

        table_number = guest.table.number if guest.table else _("non assignée")
        messages.success(request, _("Bienvenue {} ! Table {}.").format(guest.get_full_name(), table_number))

        return render(request, self.success_template, {
            'guest': guest,
            'table_number': table_number,
        })


# ============================================================
# 7. INVITATION PDF
# ============================================================

class InvitationPDFView(View):
    """Génère et télécharge l'invitation PDF."""

    def get(self, request, token):
        guest_response = get_object_or_404(GuestResponse, invitation_token=token)

        if not guest_response.submitted_at:
            messages.warning(request, _("Vous devez d'abord confirmer votre présence."))
            return redirect('guests:rsvp', token=token)

        pdf = generate_invitation_pdf(guest_response)
        if not pdf:
            messages.error(request, _("Impossible de générer l'invitation."))
            return redirect('events:event_detail', slug=guest_response.event.slug)

        filename = f"invitation_{guest_response.event.slug}_{guest_response.first_name}_{guest_response.last_name}.pdf"
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class InvitationPreviewView(View):
    """Prévisualiser l'invitation PDF dans le navigateur."""

    def get(self, request, token):
        guest_response = get_object_or_404(GuestResponse, invitation_token=token)

        if request.user.is_authenticated:
            event = guest_response.event
            is_organizer = (event.main_organizer == request.user)
            is_collaborator = event.collaborators.filter(user=request.user, status='accepted').exists()
        else:
            is_organizer = False
            is_collaborator = False

        is_guest = (request.GET.get('email') == guest_response.email)

        if not (is_organizer or is_collaborator or is_guest):
            messages.error(request, _("Accès non autorisé."))
            return redirect('authentication:login')

        pdf = generate_invitation_pdf(guest_response)
        if not pdf:
            messages.error(request, _("Impossible de générer l'invitation."))
            return redirect('events:event_detail', slug=guest_response.event.slug)

        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="invitation_{guest_response.event.slug}.pdf"'
        return response


# ============================================================
# 8. AJOUT MANUEL D'UN INVITÉ PRÉ-ENREGISTRÉ
# ============================================================

class AddInvitedGuestView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    """Ajouter manuellement un invité à la liste officielle."""
    model = InvitedGuest
    form_class = InvitedGuestForm
    template_name = 'guests/add_guest.html'

    def test_func(self):
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        return self.request.user == self.event.main_organizer

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['event'] = self.event
        return context

    def form_valid(self, form):
        form.instance.event = self.event
        form.instance.created_by = self.request.user
        messages.success(self.request, _('Invité ajouté avec succès.'))
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('guests:invited_list', kwargs={'event_id': self.event.id})


# ============================================================
# 9. LISTE DES INVITÉS PRÉ-ENREGISTRÉS (INVITED GUEST)
# ============================================================

class InvitedGuestListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """Liste des invités pré-enregistrés (InvitedGuest)."""
    model = InvitedGuest
    template_name = 'guests/invited_guest_list.html'
    context_object_name = 'invited_guests'
    paginate_by = 20

    def test_func(self):
        self.event = get_object_or_404(Event, id=self.kwargs.get('event_id'))
        user = self.request.user
        return (self.event.main_organizer == user or
                self.event.collaborators.filter(user=user, status='accepted').exists())

    def get_paginate_by(self, queryset):
        per_page = self.request.GET.get('per_page')
        if per_page and per_page.isdigit():
            per_page = int(per_page)
            if per_page in [10, 15, 20, 30, 50, 100]:
                return per_page
        return getattr(settings, 'GUESTS_PER_PAGE', 20)

    def get_queryset(self):
        queryset = InvitedGuest.objects.filter(event=self.event)
        queryset = queryset.only(
            'id', 'first_name', 'last_name', 'middle_name',
            'email', 'phone', 'created_at', 'table_id'
        ).select_related('table')

        search_query = self.request.GET.get('q', '')
        if search_query:
            queryset = queryset.filter(
                Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query) |
                Q(middle_name__icontains=search_query) |
                Q(email__icontains=search_query)
            )

        sort_by = self.request.GET.get('sort', 'last_name')
        if sort_by in ['first_name', 'last_name', 'email', 'created_at']:
            queryset = queryset.order_by(sort_by)
        else:
            queryset = queryset.order_by('last_name', 'first_name')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['event'] = self.event

        qs = InvitedGuest.objects.filter(event=self.event)
        context['total'] = qs.count()
        context['has_table'] = qs.filter(table__isnull=False).count()
        context['no_table'] = context['total'] - context['has_table']

        context['search_query'] = self.request.GET.get('q', '')
        context['current_sort'] = self.request.GET.get('sort', 'last_name')
        context['current_per_page'] = self.get_paginate_by(self.get_queryset())

        return context


# ============================================================
# 10. EXPORTS DES INVITÉS PRÉ-ENREGISTRÉS
# ============================================================

class ExportInvitedCSVView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Export CSV des invités pré-enregistrés."""

    def test_func(self):
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        return self.request.user == self.event.main_organizer

    def get(self, request, event_id):
        guests = InvitedGuest.objects.filter(event=self.event).order_by('last_name', 'first_name')
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="invited_guests_{self.event.id}.csv"'
        response.write('\ufeff')
        writer = csv.writer(response)
        writer.writerow(['Prénom', 'Nom', 'Postnom', 'Email', 'Téléphone', 'Table'])
        for g in guests:
            writer.writerow([
                g.first_name, g.last_name, g.middle_name or '',
                g.email or '', g.phone or '',
                g.table.number if g.table else ''
            ])
        return response


class ExportInvitedExcelView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Export Excel des invités pré-enregistrés."""

    def test_func(self):
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        return self.request.user == self.event.main_organizer

    def get(self, request, event_id):
        from openpyxl import Workbook
        from openpyxl.styles import Font

        guests = InvitedGuest.objects.filter(event=self.event).order_by('last_name', 'first_name')
        wb = Workbook()
        ws = wb.active
        ws.title = "Invités"
        headers = ['Prénom', 'Nom', 'Postnom', 'Email', 'Téléphone', 'Table']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).font = Font(bold=True)
        for row, g in enumerate(guests, 2):
            ws.cell(row=row, column=1, value=g.first_name)
            ws.cell(row=row, column=2, value=g.last_name)
            ws.cell(row=row, column=3, value=g.middle_name or '')
            ws.cell(row=row, column=4, value=g.email or '')
            ws.cell(row=row, column=5, value=g.phone or '')
            ws.cell(row=row, column=6, value=g.table.number if g.table else '')
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="invited_guests_{self.event.id}.xlsx"'
        wb.save(response)
        return response


# ============================================================
# 11. ASSIGNATION MANUELLE D'UNE TABLE (L)
# ============================================================

class AssignGuestTableView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Assigner manuellement un invité à une table."""

    def test_func(self):
        self.guest = get_object_or_404(GuestResponse, id=self.kwargs['guest_id'])
        return self.request.user == self.guest.event.main_organizer

    def post(self, request, guest_id):
        guest = self.guest
        table_id = request.POST.get('table_id')
        if table_id:
            try:
                table = Table.objects.get(id=table_id, event=guest.event)
                guest.table = table
                guest.save()
                messages.success(request, _('Table assignée avec succès.'))
            except Table.DoesNotExist:
                messages.error(request, _('Table invalide.'))
        else:
            guest.table = None
            guest.save()
            messages.info(request, _('Assignation retirée.'))
        return redirect(request.META.get('HTTP_REFERER', reverse('events:event_detail', kwargs={'slug': guest.event.slug})))