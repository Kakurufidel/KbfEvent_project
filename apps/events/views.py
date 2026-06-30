import logging
from django.utils import timezone 
from django.contrib.auth import login
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse, reverse_lazy
from django.utils.translation import gettext as _
from django.views.generic import CreateView, DetailView, ListView, UpdateView, View, FormView, DeleteView
from django.shortcuts import render
from django.views.generic import TemplateView
from .forms import EventForm
from apps.guests.forms import RSVPForm
from .models import Event, EventCollaborator
from django.urls import reverse_lazy
from .models import Table
from apps.guests.models import InvitedGuest
from apps.guests.forms import InvitedGuestForm
from .forms import TableForm
from apps.guests.services import TableAssignmentService
from django.http import HttpResponse
from apps.guests.models import GuestResponse
from .forms import EventForm
from .models import Event, EventCollaborator


logger = logging.getLogger(__name__)


def user_can_manage_event(user, event):
    """Vérifie si l'utilisateur peut gérer l'événement"""
    return (event.main_organizer == user or 
            EventCollaborator.objects.filter(event=event, user=user, status='accepted').exists())


class EventListView(LoginRequiredMixin, ListView):
    template_name = 'events/event_list.html'
    context_object_name = 'events'

    def get_queryset(self):
        user = self.request.user
        # Événements dont l'utilisateur est l'organisateur principal
        owned = Event.objects.filter(main_organizer=user)
        # Événements où l'utilisateur est co-organisateur (status='accepted')
        coorganized = Event.objects.filter(
            collaborators__user=user,
            collaborators__status='accepted'
        )
        # Union des deux
        return (owned | coorganized).distinct().order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['is_organizer'] = self.request.user.owned_events.exists()
        return context
    
class EventCreateView(LoginRequiredMixin, CreateView):
    """Créer un nouvel événement"""
    model = Event
    form_class = EventForm
    template_name = 'events/event_form.html'
    success_url = reverse_lazy('authentication:dashboard')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Créer un événement')
        return context

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        event = form.save(commit=False)
        event.main_organizer = self.request.user
        event.save()
        messages.success(self.request, _('Événement créé avec succès !'))
        return redirect(self.success_url)

    def form_invalid(self, form):
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(self.request, f"{error}")
        return super().form_invalid(form)


class EventUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    """Modifier un événement existant"""
    model = Event
    form_class = EventForm
    template_name = 'events/event_form.html'
    slug_url_kwarg = 'slug'
    success_url = reverse_lazy('authentication:dashboard')
    
    def test_func(self):
        event = self.get_object()
        return user_can_manage_event(self.request.user, event)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Modifier l\'événement')
        return context

    def form_valid(self, form):
        form.save()
        messages.success(self.request, _('Événement modifié avec succès !'))
        return redirect(self.success_url)


class EventDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Event
    template_name = 'events/event_detail.html'
    context_object_name = 'event'
    slug_url_kwarg = 'slug'

    def test_func(self):
        event = self.get_object()
        # Vérifier que l'utilisateur est organisateur principal ou co-organisateur accepté
        return (event.main_organizer == self.request.user or
                event.collaborators.filter(user=self.request.user, status='accepted').exists())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        event = self.object

        # Récupérer les réponses
        responses = event.responses.all()

        # Statistiques de base
        total_responses = responses.count()
        will_attend = responses.filter(will_attend=True).count()
        will_not_attend = responses.filter(will_attend=False).count()
        verified = responses.filter(verification_status='verified').count()
        unverified = responses.filter(verification_status='unverified').count()
        checkins = responses.filter(checkin_time__isnull=False).count()

        # Taux de présence (basé sur les réponses vérifiées)
        attendance_rate = 0
        if verified > 0:
            attendance_rate = round((responses.filter(will_attend=True, verification_status='verified').count() / verified) * 100, 1)

        # Nombre de personnes attendues (somme des number_of_guests pour ceux qui ont dit oui et sont vérifiés)
        expected_guests = 0
        for r in responses.filter(will_attend=True, verification_status='verified'):
            expected_guests += r.number_of_guests

        # Nombre total d'invités pré-enregistrés
        total_invited = event.invited_guests.count()

        context.update({
            'stats': {
                'total_invited': total_invited,
                'total_responses': total_responses,
                'will_attend': will_attend,
                'will_not_attend': will_not_attend,
                'verified': verified,
                'unverified': unverified,
                'checkins': checkins,
                'attendance_rate': attendance_rate,
                'expected_guests': expected_guests,
            },
            'recent_responses': responses.order_by('-submitted_at')[:10],
            'collaborators': event.collaborators.filter(status='accepted').select_related('user'),
            'rsvp_url': event.get_rsvp_url(),
            'coorganizer_url': event.get_coorganizer_url(),
        })
        return context

class EventDeleteView(LoginRequiredMixin, View):
    """Supprimer un événement"""
    def post(self, request, *args, **kwargs):
        event = get_object_or_404(Event, slug=kwargs.get('slug'), main_organizer=request.user)
        name = event.name
        event.delete()
        messages.success(request, _('L\'événement "%(name)s" a été supprimé.') % {'name': name})
        return redirect('events:event_list')


class JoinCoOrganizerView(FormView):
    """Vue pour rejoindre comme co-organisateur (basée sur classe)"""
    template_name = 'events/join_coorganizer.html'
    
    def dispatch(self, request, *args, **kwargs):
        self.event = get_object_or_404(Event, slug=kwargs.get('slug'), coorganizer_token=kwargs.get('token'))
        return super().dispatch(request, *args, **kwargs)
    
    def get_form_class(self):
        from apps.authentication.forms import RegisterForm
        return RegisterForm
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['event'] = self.event
        return context
    
    def form_valid(self, form):
        # Créer l'utilisateur
        user = form.save()
    
        # Connecter l'utilisateur
        login(self.request, user)
        
        # Ajouter comme co-organisateur
        EventCollaborator.objects.create(
            event=self.event,
            user=user,
            status='accepted',
            accepted_at=timezone.now()
        )
        
        messages.success(self.request, f'Bienvenue ! Vous êtes co-organisateur de "{self.event.name}"')
        return redirect('events:event_detail', slug=self.event.slug)
    
    def form_invalid(self, form):
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(self.request, f"{field}: {error}")
        return super().form_invalid(form)
    
    def get(self, request, *args, **kwargs):
        # Si déjà connecté, l'ajouter directement comme co-organisateur
        if request.user.is_authenticated:
            collaborator, created = EventCollaborator.objects.get_or_create(
                event=self.event,
                user=request.user,
                defaults={'status': 'accepted', 'accepted_at': timezone.now()}
            )
            if not created and collaborator.status == 'pending':
                collaborator.status = 'accepted'
                collaborator.accepted_at = timezone.now()
                collaborator.save()
            
            messages.success(request, f'Vous êtes maintenant co-organisateur de "{self.event.name}"')
            return redirect('events:event_detail', slug=self.event.slug)
        
        return super().get(request, *args, **kwargs)

class JoinCoOrganizerShortCodeView(View):
    """
    Permet à un utilisateur de rejoindre un événement comme co-organisateur
    en utilisant le code court de 6 caractères.
    """
    def get(self, request, short_code):
        try:
            event = Event.objects.get(coorganizer_short_code=short_code.upper())
        except Event.DoesNotExist:
            messages.error(request, _('Code co-organisateur invalide.'))
            return redirect('events:event_list')

        # Si l'utilisateur n'est pas connecté, le rediriger vers login avec le code
        if not request.user.is_authenticated:
            login_url = reverse('authentication:login')
            return redirect(f'{login_url}?coorganizer_code={short_code.upper()}')

        # Si déjà connecté, l'ajouter comme co-organisateur
        collaborator, created = EventCollaborator.objects.get_or_create(
            event=event,
            user=request.user,
            defaults={'status': 'accepted', 'accepted_at': timezone.now()}
        )
        if not created and collaborator.status == 'pending':
            collaborator.status = 'accepted'
            collaborator.accepted_at = timezone.now()
            collaborator.save()
        elif not created and collaborator.status == 'accepted':
            messages.info(request, _('Vous êtes déjà co-organisateur de cet événement.'))
        else:
            messages.success(request, _('Vous êtes maintenant co-organisateur de l\'événement "%s".') % event.name)

        return redirect('events:event_detail', slug=event.slug)



class RSVPFormView(FormView):
    template_name = 'guests/rsvp.html'
    form_class = RSVPForm

    def dispatch(self, request, *args, **kwargs):
        self.event = get_object_or_404(
            Event,
            slug=kwargs.get('slug'),
            rsvp_token=kwargs.get('token')
        )
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['event'] = self.event
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['event'] = self.event
        context['guest'] = None
        return context

    def form_valid(self, form):
        response = form.save(commit=False)
        response.event = self.event
        response.ip_address = self.request.META.get('REMOTE_ADDR')
        response.save()

        # Vérification automatique
        response.verify_against_invited_list()

        # Envoi email (si configuré)
        response.send_confirmation_email()

        # Message de succès
        messages.success(
            self.request,
            _('Merci ! Votre réponse a bien été enregistrée.')
        )

        # Redirection vers la page de remerciement avec les données
        return render(self.request, 'guests/rsvp_thanks.html', {
            'event': self.event,
            'response': response,
            'will_attend': response.will_attend,
        })

    def form_invalid(self, form):
        # Afficher les erreurs du formulaire
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(self.request, f"{field}: {error}")
        return super().form_invalid(form)

class TableListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Table
    template_name = 'events/table_list.html'
    context_object_name = 'tables'

    def test_func(self):
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        return self.request.user == self.event.main_organizer

    def get_queryset(self):
        return Table.objects.filter(event=self.event)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['event'] = self.event
        return context


class TableCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Table
    form_class = TableForm
    template_name = 'events/table_form.html'

    def get_initial(self):
        initial = super().get_initial()
        # Générer automatiquement le prochain numéro de table
        last_table = Table.objects.filter(event=self.event).order_by('-number').first()
        if last_table:
            try:
                next_number = int(last_table.number) + 1
                initial['number'] = str(next_number)
            except ValueError:
                initial['number'] = '1'
        else:
            initial['number'] = '1'
        return initial

class TableUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Table
    form_class = TableForm
    template_name = 'events/table_form.html'

    def test_func(self):
        table = self.get_object()
        return self.request.user == table.event.main_organizer

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['event'] = self.object.event
        context['title'] = _('Modifier la table')
        return context

    def form_valid(self, form):
        messages.success(self.request, _('Table modifiée avec succès.'))
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('events:table_list', kwargs={'event_id': self.object.event.id})


class TableDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Table
    template_name = 'events/table_confirm_delete.html'

    def test_func(self):
        table = self.get_object()
        return self.request.user == table.event.main_organizer

    def get_success_url(self):
        return reverse_lazy('events:table_list', kwargs={'event_id': self.object.event.id})

    def delete(self, request, *args, **kwargs):
        messages.success(request, _('Table supprimée avec succès.'))
        return super().delete(request, *args, **kwargs)


class AutoAssignTablesView(LoginRequiredMixin, UserPassesTestMixin, View):
    def test_func(self):
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        return self.request.user == self.event.main_organizer

    def post(self, request, event_id):
        service = TableAssignmentService(self.event)
        result = service.auto_assign_all()
        if result:
            messages.success(request, _('Les invités ont été attribués aux tables.'))
        else:
            messages.warning(request, _('Aucune table disponible ou aucun invité à attribuer.'))
        return redirect('events:event_detail', slug=self.event.slug)

class TablesPDFView(LoginRequiredMixin, UserPassesTestMixin, View):
    def test_func(self):
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        return self.request.user == self.event.main_organizer

    def get(self, request, event_id):
        from apps.guests.services import generate_tables_pdf
        pdf = generate_tables_pdf(self.event)
        if not pdf:
            messages.error(request, _('La génération du PDF a échoué. Vérifiez que reportlab est installé.'))
            return redirect('events:event_detail', slug=self.event.slug)
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="tables_{self.event.slug}.pdf"'
        return response

# class AddInvitedGuestView(LoginRequiredMixin, CreateView):
#     model = InvitedGuest
#     form_class = InvitedGuestForm
#     template_name = 'events/add_invited_guest.html'

#     def dispatch(self, request, *args, **kwargs):
#         self.event = get_object_or_404(Event, slug=kwargs.get('slug'), main_organizer=request.user)
#         return super().dispatch(request, *args, **kwargs)

#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context['event'] = self.event
#         return context

#     def form_valid(self, form):
#         form.instance.event = self.event
#         form.instance.created_by = self.request.user
#         response = super().form_valid(form)
#         messages.success(self.request, _('Invité ajouté avec succès.'))
#         return response

#     def get_success_url(self):
#         return reverse_lazy('events:event_detail', kwargs={'slug': self.event.slug})

# ========== RÉVISION MANUELLE DES TABLES ==========

class AssignGuestTableView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Vue pour déplacer un invité d'une table à une autre"""
    template_name = 'events/assign_guest_table.html'

    def test_func(self):
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        return self.request.user == self.event.main_organizer

    def get(self, request, event_id):
        guests = GuestResponse.objects.filter(
            event=self.event,
            will_attend=True
        ).select_related('table').order_by('first_name', 'last_name')
        
        tables = Table.objects.filter(event=self.event).order_by('number')
        
        context = {
            'event': self.event,
            'guests': guests,
            'tables': tables,
            'selected_guest_id': request.GET.get('guest_id'),
        }
        return render(request, self.template_name, context)

    def post(self, request, event_id):
        guest_id = request.POST.get('guest_id')
        table_id = request.POST.get('table_id')
        
        if not guest_id or not table_id:
            messages.error(request, _('Veuillez sélectionner un invité et une table.'))
            return redirect('events:assign_guest_table', event_id=event_id)
        
        try:
            guest = GuestResponse.objects.get(id=guest_id, event=self.event)
            table = Table.objects.get(id=table_id, event=self.event)
            
            # Vérifier que la table a de la place
            if table.guests.count() >= table.capacity:
                messages.error(request, _('Cette table est pleine (capacité: {capacity}).').format(capacity=table.capacity))
                return redirect('events:assign_guest_table', event_id=event_id)
            
            old_table = guest.table
            guest.table = table
            guest.save()
            
            messages.success(
                request,
                _('{guest} a été déplacé(e) de la table {old} vers la table {new}.').format(
                    guest=guest.get_full_name(),
                    old=old_table.number if old_table else 'aucune',
                    new=table.number
                )
            )
        except GuestResponse.DoesNotExist:
            messages.error(request, _('Invité introuvable.'))
        except Table.DoesNotExist:
            messages.error(request, _('Table introuvable.'))
        
        return redirect('events:assign_guest_table', event_id=event_id)
# ========== EXPORT TABLES ==========

class ExportTablesCSVView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Export CSV des tables avec invités et boissons"""
    
    def test_func(self):
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        return self.request.user == self.event.main_organizer

    def get(self, request, event_id):
        tables = Table.objects.filter(event=self.event).prefetch_related('guests')
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="tables_guests_{self.event.id}.csv"'
        response.write('\ufeff')  # BOM pour UTF-8
        
        writer = csv.writer(response)
        writer.writerow([
            'Table', 'Capacité', 'Occupée', 'Invité', 'Email', 'Téléphone',
            'Boisson', 'Autre boisson', 'Accompagnement', 'Boisson accompagnant',
            'Vérifié', 'Statut présence', 'Nombre de personnes'
        ])
        
        for table in tables:
            guests = table.guests.filter(will_attend=True).order_by('first_name', 'last_name')
            if guests.exists():
                for guest in guests:
                    writer.writerow([
                        table.number,
                        table.capacity,
                        guests.count(),
                        guest.get_full_name(),
                        guest.email,
                        guest.phone or '',
                        guest.drink_display,
                        guest.drink_other or '',
                        guest.get_companion_full_name() if guest.is_accompanied else '',
                        guest.companion_drink_display if guest.is_accompanied else '',
                        'Oui' if guest.is_verified else 'Non',
                        'Présent' if guest.will_attend else 'Absent',
                        guest.number_of_guests,
                    ])
            else:
                # Table vide
                writer.writerow([
                    table.number,
                    table.capacity,
                    0,
                    '-- Aucun invité --', '', '', '', '', '', '', '', '', ''
                ])
        
        return response


class ExportTablesExcelView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Export Excel des tables avec invités et boissons"""
    
    def test_func(self):
        self.event = get_object_or_404(Event, id=self.kwargs['event_id'])
        return self.request.user == self.event.main_organizer

    def get(self, request, event_id):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        tables = Table.objects.filter(event=self.event).prefetch_related('guests')
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Tables et invités"
        
        # En-têtes
        headers = [
            'Table', 'Capacité', 'Occupée', 'Invité', 'Email', 'Téléphone',
            'Boisson', 'Autre boisson', 'Accompagnement', 'Boisson accompagnant',
            'Vérifié', 'Statut présence', 'Nombre de personnes'
        ]
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        row_idx = 2
        for table in tables:
            guests = table.guests.filter(will_attend=True).order_by('first_name', 'last_name')
            if guests.exists():
                for guest in guests:
                    ws.cell(row=row_idx, column=1, value=table.number)
                    ws.cell(row=row_idx, column=2, value=table.capacity)
                    ws.cell(row=row_idx, column=3, value=guests.count())
                    ws.cell(row=row_idx, column=4, value=guest.get_full_name())
                    ws.cell(row=row_idx, column=5, value=guest.email)
                    ws.cell(row=row_idx, column=6, value=guest.phone or '')
                    ws.cell(row=row_idx, column=7, value=guest.drink_display)
                    ws.cell(row=row_idx, column=8, value=guest.drink_other or '')
                    ws.cell(row=row_idx, column=9, value=guest.get_companion_full_name() if guest.is_accompanied else '')
                    ws.cell(row=row_idx, column=10, value=guest.companion_drink_display if guest.is_accompanied else '')
                    ws.cell(row=row_idx, column=11, value='Oui' if guest.is_verified else 'Non')
                    ws.cell(row=row_idx, column=12, value='Présent' if guest.will_attend else 'Absent')
                    ws.cell(row=row_idx, column=13, value=guest.number_of_guests)
                    row_idx += 1
            else:
                ws.cell(row=row_idx, column=1, value=table.number)
                ws.cell(row=row_idx, column=2, value=table.capacity)
                ws.cell(row=row_idx, column=3, value=0)
                ws.cell(row=row_idx, column=4, value='-- Aucun invité --')
                row_idx += 1
        
        # Ajuster les colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="tables_guests_{self.event.id}.xlsx"'
        wb.save(response)
        return response
    

class CollaboratorScanPermissionView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Basculer le droit de scan d'un co-organisateur"""
    
    def test_func(self):
        collaborator = get_object_or_404(EventCollaborator, id=self.kwargs['pk'])
        return self.request.user == collaborator.event.main_organizer
    
    def post(self, request, pk):
        collaborator = get_object_or_404(EventCollaborator, id=pk)
        collaborator.can_scan = not collaborator.can_scan
        collaborator.save()
        messages.success(request, _('Droit de scan modifié.'))
        return redirect('events:event_detail', slug=collaborator.event.slug)
    
class CollaboratorDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    """Supprimer un co-organisateur"""
    model = EventCollaborator
    template_name = 'events/collaborator_confirm_delete.html'
    
    def test_func(self):
        collaborator = self.get_object()
        return self.request.user == collaborator.event.main_organizer
    
    def get_success_url(self):
        return reverse_lazy('events:event_detail', kwargs={'slug': self.object.event.slug})
    
    def delete(self, request, *args, **kwargs):
        collaborator = self.get_object()
        messages.success(request, _('Co-organisateur retiré avec succès.'))
        return super().delete(request, *args, **kwargs)